from calendar import weekday
from email import message
from fileinput import filename
from openpyxl import Workbook
import requests
requests.packages.urllib3.disable_warnings()
import os
import datetime
from requests_kerberos import HTTPKerberosAuth, REQUIRED, OPTIONAL
import openpyxl
from openpyxl.workbook.protection import WorkbookProtection
import time
import datetime as dt
from datetime import timedelta
from datetime import date
import datetime
from tkinter import messagebox




def testbrowser():
    try:
        url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(82c69b17-7c44-48c1-8374-7f4850747021)'
        response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
    except Exception as e:
        messagebox.showwarning("Error", "Ensure SIM website is opened and accessed in your browser \nOnce done, rerun program \n\n"+ str(e.__class__))
        quit()
    
def initializeSWAT():
    try: 
        print("\nGetting SIMs information...\n")
        tasklistlocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStrSWAT + '\\shift_config.xlsx'
        wb = openpyxl.load_workbook(tasklistlocation)
    except Exception as e:
        messagebox.showwarning("Error", "File shift_config_xlsx in CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\" + dateStrSWAT +" couldnt be found \nMake sure such directory in share drive exists \n\n"+ str(e.__class__))
        quit()

    try:
        schedulernamelist=[]
        ws = wb['ExtraTasks']
        range = ws.iter_rows()
        for row in range:
            for cell in row:
                if (cell.value == "PM Shift Evening Compliance"):
                    for number in [2,3,4,5,6,7,8,9,10]:
                        if ws.cell(row=cell.row, column=number).value != None:
                            schedulernamelist.append(ws.cell(row=cell.row, column=number).value)
                            number+=1
        schedulernameliststring = "+OR+".join(schedulernamelist)
    
        global DPOreviewerstring
        DPOreviewer = []
        ws = wb['ExtraTasks']
        range = ws.iter_rows()
        for row in range:
            for cell in row:
                if (cell.value == "DSP Config Review "):
                    for number in [2,3,4,5,6,7,8,9,10]:
                        if ws.cell(row=cell.row, column=number).value != None:
                            DPOreviewer.append(ws.cell(row=cell.row, column=number).value)
                            number+=1
        DPOreviewerstring = ", ".join(DPOreviewer)
    except Exception as e:
        messagebox.showwarning("Error", "Schedulers logins couldnt be retrieved from tab ExtraTasks in File shift_config_xlsx in CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\" + dateStrSWAT +"\nMake sure the file has the same strucutre as days before and no one has modified \n\n"+ str(e.__class__))
        quit()

    try:
        #Manual Changes
        numberofSIMsManualChanges=[]
        urlManualChanges = 'https://issues.amazon.com/issues/search?q=containingFolder%3A(82c69b17-7c44-48c1-8374-7f4850747021)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-12HOURS..NOW%5D)+assignee%3A('+schedulernameliststring+')&sort=lastUpdatedDate+desc&selectedDocument=71d887b3-6485-4176-b4a1-78bd8ef0b113'
        count = 0
        for schedulername in schedulernamelist:
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(82c69b17-7c44-48c1-8374-7f4850747021)%20assignee:(' +schedulername+ ')%20status:(Resolved)%20lastResolvedDate:[NOW-12HOURS%20TO%20NOW]&sort=lastUpdatedConversationDate%20desc&selectedDocument=1f923d52-9dc4-4351-a733-44d5ede9d152'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            numberofSIMsManualChanges.append((response['totalNumberFound']))
            count += 1
            print("Retrieving information Manual Changes SIMs: " + str(((int(count)/int(len(schedulernamelist)))*100))[0:5] + "%", end="\r")
        print("------Success: Manual Changes SIMs retrieved------")
    
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read Manual Changes SIMs, please contact BIA team \n\n"+ str(e.__class__))
        quit()
    
        
    
    try:
        #Emergency Adhoc SIMs
        count1=0
        numberofSIMsemergency =[]
        urlnumberofSIMsemergency = 'https://issues.amazon.com/issues/search?q=containingFolder%3A(8af3e156-8682-48bb-bf51-3beef6455e94)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-12HOURS..NOW%5D)+assignee%3A('+schedulernameliststring+')&sort=lastUpdatedDate+desc&selectedDocument=1d85a7a9-379d-40ae-98e7-0b214cf3a76a'
        for schedulername in schedulernamelist:
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(8af3e156-8682-48bb-bf51-3beef6455e94)%20assignee:(' +schedulername+ ')%20status:(Resolved)%20lastResolvedDate:[NOW-12HOURS%20TO%20NOW]&sort=lastUpdatedConversationDate%20desc&selectedDocument=1f923d52-9dc4-4351-a733-44d5ede9d152'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            numberofSIMsemergency.append((response['totalNumberFound']))
            count1 += 1
            print("Retrieving information Emergency Status SIMs: " + str(((int(count1)/int(len(schedulernamelist)))*100))[0:5] + "%", end="\r")
        print("------Success: Emergency Status SIMs retrieved------")
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read Emergency SIMs, please contact BIA team \n\n"+ str(e.__class__))
        quit()
        
    
    try:
        #DPO SIMs
        count2 = 0
        numberofSIMsDPO = []
        urlnumberofSIMsDPO = 'https://issues.amazon.com/issues/search?q=conversation.author%3A('+ schedulernameliststring +')+containingFolder%3A(daee11b4-0519-45e8-a29a-cece4907bc30)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-12HOURS..NOW%5D)&sort=lastUpdatedDate+desc&selectedDocument=c5dd1220-1c39-4773-bc2e-53cf40359cfe'
        for schedulername in schedulernamelist:
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(44f5c924-891b-44ca-9699-394593062104%20OR%20daee11b4-0519-45e8-a29a-cece4907bc30)%20conversation.author:('+schedulername+')%20status:(Resolved)%20lastResolvedDate:[NOW-12HOURS%20TO%20NOW]&sort=lastUpdatedConversationDate%20desc&selectedDocument=802ebc9c-a964-4792-b2d8-7fbba751912d'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            numberofSIMsDPO.append((response['totalNumberFound']))
            count2 += 1
            print("Retrieving information DPO SIMs: " + str(((int(count2)/int(len(schedulernamelist)))*100))[0:5] + "%", end="\r")
        print("------Success: DPO SIMs retrieved------")
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read DPO SIMs, please contact BIA team \n\n"+ str(e.__class__))
        quit()

        ##formatting message
    try:
        datatable = ""
        for scheduler,numberSIMs,numberemergency,numberDPO in zip(schedulernamelist, numberofSIMsManualChanges,numberofSIMsemergency,numberofSIMsDPO):
            datatable += "\n|" + str(scheduler) + "|" + str(numberSIMs)+ "|" + str(numberemergency) + "|" + str(numberDPO) + "|" + str(numberSIMs+numberemergency+numberDPO)
        
        messageprint = "\nSending webhooks to Chime Rooms SWAT..."
        data = {"Content": "/md \n**Number of SIMs per scheduler:**\n\n|Scheduler|[Count Manual Changes]("+ urlManualChanges +")|[Count Status Change]("+ urlnumberofSIMsemergency +")|[Count DPO SIMs]("+ urlnumberofSIMsDPO +")|**Total**|\n|---|---|---|---|"+ datatable +"\nScheduler *"+ DPOreviewerstring + "* reviewed DPO SIMs, therefore the scheduler is likely to have a higher number in DPO column \nDPO SIMs column shows the number of SIMs where scheduler has either attached CSV / updated DPO / reviewed DPO"}
        sendwebhook(data,messageprint) #sending webhook
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't properly format the tables, please contact BIA team \n\n"+ str(e.__class__))
        quit()

    try:
        #getting all the DS nodes
        DSlist = []
        wDS = wb['Stations']
        for cell in wDS["A"]:
            DSlist.append(cell.value)
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't retrieved all DS nodes from tab Stations in File shift_config_xlsx in CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\" + dateStrSWAT +"\nMake sure the file has the same strucutre as days before and no one has modified \n\n"+ str(e.__class__))
        quit()
    
    try:
        countDS=0
        numberofSIMsDS =[]
        DStopoffenders = []
        DSlink = []
        for DSnode in DSlist:
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(82c69b17-7c44-48c1-8374-7f4850747021) status:(Resolved) lastResolvedDate:[NOW-12HOURS TO NOW] ('+ DSnode + ') &sort=lastUpdatedConversationDate desc&selectedDocument=1f923d52-9dc4-4351-a733-44d5ede9d152'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            linkSIM = "https://issues.amazon.com/issues/search?q=status%3A(Resolved)+lastResolvedDate%3A(%5BNOW-12HOURS..NOW%5D)+("+ DSnode +")+containingFolder%3A(82c69b17-7c44-48c1-8374-7f4850747021)&sort=lastUpdatedDate+desc&selectedDocument=802ebc9c-a964-4792-b2d8-7fbba751912d"
            countDS += 1
            print("Retrieving information DS raising Manual Changes SIMs: " + str(((int(countDS)/int(len(DSlist)))*100))[0:5] + "%", end="\r")
            if int((response['totalNumberFound']))>3:
                numberofSIMsDS.append((response['totalNumberFound']))
                DStopoffenders.append(DSnode)
                DSlink.append(linkSIM)
        print("------Success: DS SIMs for Manual Changes retrieved------")

        if len(DStopoffenders) == 0:
            DStopoffenders.append("-")
            numberofSIMsDS.append("No DS has raised more than 3 Manual Changes SIMs")
            DSlink.append('https://sim.amazon.com/issues/search?q=status%3A(Open)+containingFolder%3A(82c69b17-7c44-48c1-8374-7f4850747021)&sort=lastUpdatedConversationDate+desc&selectedDocument=22b60b16-cea4-40a7-a82d-20400baa9301')
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read how many Manual Changes SIMs has been raised by every DS, please contact BIA team \n\n"+ str(e.__class__))
        quit()
    
    try:
        datatable = ""
        for DSnodewebhook, numberofSIMsDS, SIMlink in zip(DStopoffenders, numberofSIMsDS,DSlink):
            datatable += "\n|" + str(DSnodewebhook) + "|" + str(numberofSIMsDS) + "|["+DSnodewebhook+" SIM links](" + str(SIMlink) + ")|"
        
        messageprint = "\nSending webhooks to Chime Rooms SWAT..."
        data = {"Content": "/md \n**Number of SIMs per DS (only DS that raised 4 or more SIMs): **\n\n|DS|Count Manual Changes|SIMs link|\n|---|---|---|"+ datatable }
        
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't properly format the tables, please contact BIA team \n\n"+ str(e.__class__))
        quit()
    
    sendwebhook(data,messageprint) #sending webhook
    logdata(dateStrSWAT + ' SWAT logs.xlsx')

    
def logdata(filename):
    try:
        print("Logging data...")
        timenow = datetime.datetime.now().strftime('%H:%M:%S')
        logslocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs\\'  + filename
        #new workbook
        wlogs = Workbook()
        wlogssheet = wlogs.active
        
        wlogssheet.cell(row=1, column=1).value = "Logs type"
        wlogssheet.cell(row=1, column=2).value = "Time Retrieved"
        
        wlogssheet.cell(row=2, column=1).value = "SWAT Information"
        wlogssheet.cell(row=2, column=2).value = timenow
        
        wlogs.save(logslocation)
        
        print("\nLogs saved")
    
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module logs couldn't be saved today under CentralOPS\PM Shift\DHP1\SchedulingCompliance\\" + today + " Compliance Uploads logs, please contact BIA team \n\n"+ str(e.__class__))
        quit()

    

    

def sendwebhook(data,messageprint):
    result = False
    try:
        print(messageprint)
        urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/b45d5ab0-f18d-4de3-bf61-e4105b201c3d?token=MDhITUIzbUV8MXxtM3dOVlNkLXphc0QwX2kxSmxwaHV1b2NZdGZjaXJ6UVJYMkN5RWhLT2U0"
        #  test urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/11bfd17d-d0fc-4129-be49-c9dee26607f0?token=aXF6OUZLTDV8MXx6UUhZSFFJanZ6alBNWmtpamQtMDh5eC0wTmJtTHl5NlZ3WjIyeWZySUdF"
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module webhook couldn't be sent, please contact BIA team \n\n"+ str(e.__class__))
        return result
    
    
def weeklyreport():
    dateStrSWATreportlist = [(dateVarSWAT-timedelta(days=7)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=6)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=5)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=4)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=3)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=2)).strftime('%d-%m'),(dateVarSWAT-timedelta(days=1)).strftime('%d-%m')]
    daysweekinhoursstart = [19,43,67,91,115,139,163]
    daysweekinhoursend = [43,67,91,115,139,163,187]
    try:
        countweeklymanualchanges=0
        numberofSIMsmanualchangesweekly =[]
        SIMsmanualchangesweeklylinks =[]
        for hourstart,hourend in zip(daysweekinhoursstart,daysweekinhoursend):
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(82c69b17-7c44-48c1-8374-7f4850747021)%20status:(Resolved)%20lastResolvedDate:[NOW-'+str(hourend) + 'HOURS%20TO%20NOW-'+str(hourstart) + 'HOURS]&sort=lastUpdatedConversationDate%20desc&selectedDocument=2ba73a99-db81-46d0-826b-036a8d871322'
            urlnumberofSIMsmanual = 'https://issues.amazon.com/issues/search?q=containingFolder%3A(82c69b17-7c44-48c1-8374-7f4850747021)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-'+str(hourend) + 'HOURS..NOW-'+str(hourstart) + 'HOURS%5D)'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            SIMsmanualchangesweeklylinks.append((urlnumberofSIMsmanual))
            numberofSIMsmanualchangesweekly.append((response['totalNumberFound']))
            countweeklymanualchanges += 1
            print("Retrieving information Manual Changes SIMs for weekly report: " + str(((int(countweeklymanualchanges)/int(len(daysweekinhoursstart)))*100))[0:5] + "%", end="\r")
        print("----------------Success: Manual Changes SIMs retrieved----------------")
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read Manual Changes SIMs for the weekly report, please contact BIA team \n\n"+ str(e.__class__))
        quit()

    try:
        countweeklyemergency=0
        numberofSIMsemergency =[]
        SIMsemergencyweeklylinks =[]
        for hourstart,hourend in zip(daysweekinhoursstart,daysweekinhoursend):
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(8af3e156-8682-48bb-bf51-3beef6455e94)%20status:(Resolved)%20lastResolvedDate:[NOW-'+str(hourend) + 'HOURS%20TO%20NOW-'+str(hourstart) + 'HOURS]&sort=lastUpdatedConversationDate%20desc&selectedDocument=2ba73a99-db81-46d0-826b-036a8d871322'
            urlnumberofSIMs = 'https://issues.amazon.com/issues/search?q=containingFolder%3A(8af3e156-8682-48bb-bf51-3beef6455e94)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-'+str(hourend) + 'HOURS..NOW-'+str(hourstart) + 'HOURS%5D)'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            SIMsemergencyweeklylinks.append((urlnumberofSIMs))
            numberofSIMsemergency.append((response['totalNumberFound']))
            countweeklyemergency += 1
            print("Retrieving information Emergency SIMs for weekly report: " + str(((int(countweeklyemergency)/int(len(daysweekinhoursstart)))*100))[0:5] + "%", end="\r")
        print("----------------Success: Emergency SIMs retrieved-----------------")
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read Emergency SIMs for the weekly report, please contact BIA team \n\n"+ str(e.__class__))
        quit()

    try:
        countweeklyDPO=0
        numberofSIMsDPOweekly =[]
        SIMsDPOweeklylinks =[]
        for hourstart,hourend in zip(daysweekinhoursstart,daysweekinhoursend):
            url = 'https://maxis-service-prod-pdx.amazon.com/issues?q=containingFolder:(daee11b4-0519-45e8-a29a-cece4907bc30)%20status:(Resolved)%20lastResolvedDate:[NOW-'+str(hourend) + 'HOURS%20TO%20NOW-'+str(hourstart) + 'HOURS]&sort=lastUpdatedConversationDate%20desc&selectedDocument=2ba73a99-db81-46d0-826b-036a8d871322'
            urlnumberofSIMs = 'https://issues.amazon.com/issues/search?q=containingFolder%3A(daee11b4-0519-45e8-a29a-cece4907bc30)+status%3A(Resolved)+lastUpdatedConversationDate%3A(%5BNOW-'+str(hourend) + 'HOURS..NOW-'+str(hourstart) + 'HOURS%5D)'
            response = requests.get(url, auth=HTTPKerberosAuth(mutual_authentication=OPTIONAL),verify=False,allow_redirects=True,timeout=30).json()
            SIMsDPOweeklylinks.append((urlnumberofSIMs))
            numberofSIMsDPOweekly.append((response['totalNumberFound']))
            countweeklyDPO += 1
            print("Retrieving information DPO SIMs for weekly report: " + str(((int(countweeklyDPO)/int(len(daysweekinhoursstart)))*100))[0:5] + "%", end="\r")
        print("---------------Success: DPO SIMs retrieved----------------")
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't read DPO SIMs for the weekly report, please contact BIA team \n\n"+ str(e.__class__))
        quit()



    try:
        datatable = ""
        
        datatable += "\n|Manual Changes SIMs|" + str(numberofSIMsmanualchangesweekly[6]) + "|" + str(numberofSIMsmanualchangesweekly[5]) + "|"+ str(numberofSIMsmanualchangesweekly[4]) + "|"+ str(numberofSIMsmanualchangesweekly[3]) + "|"+ str(numberofSIMsmanualchangesweekly[2]) + "|"+ str(numberofSIMsmanualchangesweekly[1]) + "|"+ str(numberofSIMsmanualchangesweekly[0]) + "|"
        datatable += "\n|Emerency SIMs|"+ str(numberofSIMsemergency[6]) + "|" + str(numberofSIMsemergency[5]) + "|"+ str(numberofSIMsemergency[4]) + "|"+ str(numberofSIMsemergency[3]) + "|"+ str(numberofSIMsemergency[2]) + "|"+ str(numberofSIMsemergency[1]) + "|"+ str(numberofSIMsemergency[0]) + "|"
        datatable += "\n|DPO SIMs|"+ str(numberofSIMsDPOweekly[6]) + "|" + str(numberofSIMsDPOweekly[5]) + "|"+ str(numberofSIMsDPOweekly[4]) + "|"+ str(numberofSIMsDPOweekly[3]) + "|"+ str(numberofSIMsDPOweekly[2]) + "|"+ str(numberofSIMsDPOweekly[1]) + "|"+ str(numberofSIMsDPOweekly[0]) + "|"
        datatable += "\n|**Total SIMs**|"+ str(numberofSIMsmanualchangesweekly[6]+numberofSIMsemergency[6]+numberofSIMsDPOweekly[6]) + "|" + str(numberofSIMsmanualchangesweekly[5]+numberofSIMsemergency[5]+numberofSIMsDPOweekly[5]) + "|"+ str(numberofSIMsmanualchangesweekly[4]+numberofSIMsemergency[4]+numberofSIMsDPOweekly[4]) + "|"+ str(numberofSIMsmanualchangesweekly[3]+numberofSIMsemergency[3]+numberofSIMsDPOweekly[3]) + "|"+ str(numberofSIMsmanualchangesweekly[2]+numberofSIMsemergency[2]+numberofSIMsDPOweekly[2]) + "|"+ str(numberofSIMsmanualchangesweekly[1]+numberofSIMsemergency[1]+numberofSIMsDPOweekly[1]) + "|"+ str(numberofSIMsmanualchangesweekly[0]+numberofSIMsemergency[0]+numberofSIMsDPOweekly[0]) + "|"
        
        messageprint = "\nSending webhooks to Chime Rooms SWAT..."
        data = {"Content": "/md \n**Weekly Report (Number of SIMs per day): **\n\n|Type|Saturday ("+ str(dateStrSWATreportlist[0]) +") |Sunday ("+ str(dateStrSWATreportlist[1]) +") |Monday ("+ str(dateStrSWATreportlist[2]) +") |Tuesday ("+ str(dateStrSWATreportlist[3]) +") |Wednesday ("+ str(dateStrSWATreportlist[4]) +") |Thursday ("+ str(dateStrSWATreportlist[5]) +") |Friday ("+ str(dateStrSWATreportlist[6]) +") |\n|---|---|---|---|---|---|---|---|"+ datatable }
    except Exception as e:
        messagebox.showwarning("Error", "SWAT module couldn't properly format the tables, please contact BIA team \n\n"+ str(e.__class__))
        quit()
    
    sendwebhook(data,messageprint) #sending webhook
    logdata(dateStrSWAT + ' SWAT Weekly Report logs.xlsx')
    
    






def SWATmain():
    testbrowser()
    try:
        global dateStrSWAT
        global dateVarSWAT
        global today
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        dateVarSWAT = datetime.datetime.today()
        dateStrSWAT = dateVarSWAT.strftime('%Y-%m-%d')
        timelimitSWAT = dateVarSWAT.strptime("20:40:00",'%H:%M:%S').time()
        timelimitSWATweeklyreport = dateVarSWAT.strptime("19:00:00",'%H:%M:%S').time()
        weekday = datetime.datetime.today().weekday()

        fileexistsSWAT = os.path.exists(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs\\' + dateStrSWAT + ' SWAT logs.xlsx')
        fileexistsSWATweeklyreport = os.path.exists(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\SchedulingCompliance\\' + today + ' Compliance Uploads logs\\' + dateStrSWAT + ' SWAT Weekly Report logs.xlsx')
    except Exception as e:
        messagebox.showwarning("Error", "Problem with one of the variables defined, please contact BIA team \n\n"+ str(e.__class__))
        quit()
    
    #for daily report
    if dateVarSWAT.time() > timelimitSWAT:
        if fileexistsSWAT == False:
            print("Initializing SWAT module. . . . . .")
            initializeSWAT()
        if fileexistsSWAT == True:
            print("SWAT module has been already ran")
    
    #for weekly report (on Saturdays)
    if weekday == 5 and dateVarSWAT.time() > timelimitSWATweeklyreport:
        if fileexistsSWATweeklyreport == False:
            print("Initializing SWAT module weekly report. . . . . .")
            weeklyreport()
        if fileexistsSWATweeklyreport == True:
            pass
            


# if __name__ == "__main__":
#     SWATmain()
    

